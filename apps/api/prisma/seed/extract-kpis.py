#!/usr/bin/env python3
"""
Extract all 46 KPIs from the source spreadsheet into kpi-data.json.

Run once (and whenever the spreadsheet changes):
    python3 prisma/seed/extract-kpis.py

The resulting JSON is consumed by prisma/seed/seed.ts.
"""
from __future__ import annotations
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

import openpyxl

SPREADSHEET = Path(
    "/Users/saanvivishal/Desktop/SCORE_CARD_KPI_UPDATED.xlsx"
)
OUT = Path(__file__).parent / "kpi-data.json"

# ------------------------------------------------------------
# Condition parser — turns tier strings into structured JSONB
# ------------------------------------------------------------
#
# Tier strings come in a handful of forms:
#   "100% → 200"                      eq 100 (pct)
#   "75-99% → 150"                    range 75..99 (pct)
#   "≤25% → 0", "≤5% → 200"           lte 25 / lte 5
#   "<1 hr → 75"                      lt 1
#   ">1 & <12 hrs → 50"               range 1..12 (exclusive both)
#   ">48 hrs → 0"                     gt 48
#   ">90 days → 100"                  gt 90
#   "0/none → 50"                     eq 0
#   "1-2/mon → 40"                    range 1..2
#   ">10/mon → 0"                     gt 10
#   ">50-90% of peers → 50"           range 50..90 (approx)
#   "In seconds → 100"                eq "In seconds"
#   "Declining → 25"                  eq "Declining"
#   "Yes, contained → 50"             eq "Yes, contained"
#   "Overrun <25% → 25"               eq "Overrun <25%"  (keep the label verbatim)
#
# We preserve the raw tier label always; `condition` is a best-effort
# structured representation. The scoring engine uses condition + input
# type to bucket user input.

ARROW = re.compile(r"\s*(?:→|->|=>|:)\s*")

NUMERIC_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    # "≥95%"
    (re.compile(r"^≥\s*(\d+(?:\.\d+)?)\s*%?"), "gte"),
    # "≤25%", "≤5%"
    (re.compile(r"^[≤<=]{1,2}\s*(\d+(?:\.\d+)?)\s*%?"), "lte_or_lt"),
    # "<1 hr", "<10 days"
    (re.compile(r"^<\s*(\d+(?:\.\d+)?)"), "lt"),
    # ">10/mon", ">25%", ">48 hrs", ">90 days"
    (re.compile(r"^>\s*(\d+(?:\.\d+)?)"), "gt"),
    # ">1 & <12 hrs" / ">60 & <90 days"
    (re.compile(r"^>\s*(\d+)\s*&\s*<\s*(\d+)"), "range_ex"),
    # "75-99%", "1-2/mon", "50-64%"
    (re.compile(r"^(\d+)\s*[-–]\s*(\d+)\s*%?"), "range"),
    # "100%" / "100" / "0/none"
    (re.compile(r"^(\d+(?:\.\d+)?)\s*(?:%|/\s*none)?$"), "eq_num"),
]


def parse_tier(tier_str: str) -> dict[str, Any] | None:
    """Return {label, score, condition} or None for empty/placeholder."""
    if tier_str is None:
        return None
    raw = str(tier_str).strip()
    if raw in ("", "-", "—"):
        return None

    parts = ARROW.split(raw, maxsplit=1)
    if len(parts) != 2:
        return None
    label_part, score_part = parts[0].strip(), parts[1].strip()

    try:
        score = int(score_part.replace(",", ""))
    except ValueError:
        # Try float
        score = int(float(score_part))

    cond = _structure_condition(label_part)
    return {
        "tierLabel": label_part,
        "scoreValue": score,
        "condition": cond,
    }


def _structure_condition(label: str) -> dict[str, Any]:
    """Best-effort structured condition for a tier label."""
    s = label.strip()

    # Handle ">1 & <12 hrs" first — otherwise single-operand patterns match.
    m = re.match(r"^>\s*(\d+)\s*&\s*<\s*(\d+)", s)
    if m:
        lo = int(m.group(1))
        hi = int(m.group(2))
        return {"op": "range", "min": lo, "max": hi, "exclusive": True}

    m = re.match(r"^≥\s*(\d+(?:\.\d+)?)\s*%?", s)
    if m:
        return {"op": "gte", "value": float(m.group(1))}

    m = re.match(r"^[≤]\s*(\d+(?:\.\d+)?)\s*%?", s)
    if m:
        return {"op": "lte", "value": float(m.group(1))}

    m = re.match(r"^<\s*(\d+(?:\.\d+)?)", s)
    if m:
        return {"op": "lt", "value": float(m.group(1))}

    m = re.match(r"^>\s*(\d+(?:\.\d+)?)", s)
    if m:
        return {"op": "gt", "value": float(m.group(1))}

    m = re.match(r"^(\d+)\s*[-–]\s*(\d+)\s*%?", s)
    if m:
        return {"op": "range", "min": int(m.group(1)), "max": int(m.group(2))}

    m = re.match(r"^(\d+(?:\.\d+)?)\s*(?:%|/\s*none)?$", s)
    if m:
        return {"op": "eq", "value": float(m.group(1))}

    # "0/none", "1-2/mon" already handled above by the numeric patterns.
    # Everything else — "Declining", "Yes, contained", "In seconds",
    # "Overrun <25%", "100% in line", etc. — is a categorical string.
    return {"op": "eq", "value": s}


# ------------------------------------------------------------
# Input type inference
# ------------------------------------------------------------
#
# PERCENTAGE: all 5 tiers (minus empty) describe a 0..100 percentage bucket.
# DROPDOWN:   discrete bucket labels like "<1 hr", "In minutes", "0/none".
# RADIO:      small option sets like trend (Declining/Steady/Increasing)
#             or root cause (Yes, contained / Yes, not contained / Not sure).
#
# Heuristic: if every non-empty tier matches a pure percentage form
# ("100%", "75-99%", "≤25%", "≥95%"), treat as PERCENTAGE. Else if labels
# look like free-text choices, classify as RADIO/DROPDOWN.

PCT_LABEL = re.compile(r"^[≤≥<>]*\s*\d+(?:\s*[-–]\s*\d+)?\s*%")
TIME_LABEL = re.compile(r"\b(hr|hrs|hour|day|days|week|weeks|second|minute)\b", re.I)
COUNT_LABEL = re.compile(r"/\s*(mon|month|none)", re.I)


def infer_input_type(tier_labels: list[str]) -> str:
    nonempty = [t for t in tier_labels if t and t not in ("-", "—")]
    if not nonempty:
        return "RADIO"
    if all(PCT_LABEL.match(t) for t in nonempty):
        return "PERCENTAGE"
    if any(TIME_LABEL.search(t) for t in nonempty):
        return "DROPDOWN"
    if any(COUNT_LABEL.search(t) for t in nonempty):
        return "DROPDOWN"
    # Trend / root cause / AI adoption etc.
    return "RADIO"


# ------------------------------------------------------------
# Info text — one-line plain English for the (i) icon
# ------------------------------------------------------------
INFO_TEXT = {
    "User Awareness Training Completion": "Percentage of employees who have completed mandatory security awareness training.",
    "Phishing Click Rate": "Percentage of employees who clicked on a simulated phishing email. Lower is better.",
    "Security Awareness — Social Engineering Failure Rate": "Percentage of employees who failed a social engineering simulation. Lower is better.",
    "Security Awareness — Quiz Score (% employees passing)": "Percentage of employees who passed the security awareness quiz.",
    "Policy Violations — Data in Personal Space": "Number of times per month an employee stored company data in personal cloud, email, or device.",
    "Policy Violations — Data Classification": "Number of times per month data was mis-labelled (e.g. confidential marked public).",
    "Policy Violations — Unauthorised Software Installs": "Number of unauthorised software installs detected on company devices per month.",
    "Policy Violations — Removable Media Usage": "Number of times removable media (USB, external drives) was used against policy per month.",
    "MTTD — Avg Detection Time (weekly)": "Mean time to detect a security incident, measured weekly. Lower is better.",
    "MTTD — Month-on-Month Trend": "Direction of MTTD over the last month. Declining (getting faster) is best.",
    "MTTR — Avg Response Time (weekly)": "Mean time to respond to a detected incident, measured weekly. Lower is better.",
    "MTTR — Month-on-Month Trend": "Direction of MTTR over the last month. Declining (getting faster) is best.",
    "MTBF — Mean Time Between Failures": "Average time between security incidents. Higher is better.",
    "MTTC — Mean Time to Contain": "Average time from detection to full containment. Faster is better.",
    "Patch Compliance Rate": "Percentage of in-scope systems with all required patches applied.",
    "Security Patch Deployment Time": "Time taken to deploy a critical security patch after release.",
    "Intrusion — Malicious Login Attempts Blocked": "Percentage of malicious login attempts that were successfully blocked.",
    "Intrusion — Malicious Port Scans Blocked": "Percentage of malicious port scans that were successfully blocked.",
    "Intrusion — Known Exploit Payloads Blocked": "Percentage of known exploit payloads that were successfully blocked.",
    "Intrusion — Viruses / Ransomware Blocked": "Percentage of virus and ransomware attempts that were successfully blocked.",
    "Intrusion — False Positive Verification & Whitelist": "Percentage of false positives correctly verified and whitelisted.",
    "Zero-Day Exploit Cases — % Contained": "Percentage of zero-day exploit attempts that were contained.",
    "Data Exfiltration Attempts — % Contained": "Percentage of data exfiltration attempts that were contained.",
    "DNS & C2 Traffic — % Identified & Prevented": "Percentage of malicious DNS and command-and-control traffic identified and blocked.",
    "User Privilege Escalation — % Identified & Prevented": "Percentage of privilege escalation attempts identified and prevented.",
    "Severity of Incidents (High & Critical / month)": "Number of high and critical severity incidents per month.",
    "Root Cause — Misconfigurations": "Were misconfigurations the root cause, and were they contained?",
    "Root Cause — Phishing": "Was phishing the root cause, and was the incident contained?",
    "Root Cause — Outdated Software": "Was outdated software the root cause, and was the incident contained?",
    "Root Cause — Privileged Misuse": "Was privileged access misuse the root cause, and was the incident contained?",
    "Root Cause — Other Issues": "Were other root causes identified, and were the incidents contained?",
    "Failed Login Rate (weekly % of total logins)": "Percentage of failed logins out of total login attempts, weekly.",
    "Vulnerability Recurrence Rate (monthly %)": "Percentage of vulnerabilities that recurred after being closed.",
    "EDR Coverage — % Systems Covered": "Percentage of endpoints with EDR (endpoint detection and response) agent installed.",
    "System Hardening Status (CIS Benchmark %)": "Percentage compliance with CIS hardening benchmarks across systems.",
    "Anti-Phishing / Email Gateway Effectiveness": "Percentage of phishing emails blocked by the email gateway.",
    "Browser & Application Patch Level": "Percentage of endpoints with browsers and apps fully patched.",
    "Backup & Recovery Efficacy (Recovery Time)": "Time to fully restore systems from the most recent backup.",
    "Privileged Account Monitoring (% monitored)": "Percentage of privileged accounts under active monitoring.",
    "Unresolved Critical Vulnerabilities (avg monthly CVEs)": "Average number of unresolved critical CVEs per month.",
    "Third-Party Vendor Risk Score (monthly % adherence)": "Percentage of third-party vendors meeting security requirements.",
    "Cloud Misconfiguration Rate": "Percentage of cloud resources with security misconfigurations.",
    "Security Assessment Completion Rate": "Percentage of scheduled security assessments completed on time.",
    "Average Incident Cost (M2M trend)": "Direction of average incident cost month-on-month. Declining is best.",
    "Budget vs Actual Spend": "How closely actual security spend tracked the approved budget.",
    "AI Impact in Security (vs peer adoption)": "How your AI-in-security adoption compares to industry peers.",
}


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------
@dataclass
class TierOut:
    tierLabel: str
    scoreValue: int
    condition: dict[str, Any]
    tierOrder: int


@dataclass
class KpiOut:
    kpiName: str
    level: str
    category: str
    maxScore: int
    weightage: float
    inputType: str
    infoText: str
    nistControlIds: list[str] = field(default_factory=list)
    isoControlIds: list[str] = field(default_factory=list)
    tiers: list[TierOut] = field(default_factory=list)


LEVEL_MAP = {
    "INDIVIDUAL": "PEOPLE",
    "Individual": "PEOPLE",
    "PEOPLE": "PEOPLE",
    "PROCESS": "PROCESS",
    "Process": "PROCESS",
    "COMPANY": "COMPANY",
    "Company": "COMPANY",
}

NIST_MAP = {
    "PROTECT (PR)": ["PR.AA-01"],
    "DETECT (DE)": ["DE.CM-01"],
    "RESPOND (RS)": ["RS.AN-01"],
    "RECOVER (RC)": ["RC.RP-01"],
    "GOVERN (GV)": ["GV.OC-01"],
    "IDENTIFY (ID)": ["ID.AM-01"],
}


def main() -> int:
    if not SPREADSHEET.exists():
        print(f"Spreadsheet not found: {SPREADSHEET}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)

    # Sheet 1: scoring reference has the authoritative tier data.
    ref = wb["KPI Scoring Reference"]
    # Sheet 2: 3-level app structure has NIST / ISO mappings.
    app = wb["3-Level App Structure"]

    # Index NIST / ISO by KPI name (they share names with the scoring ref).
    nist_by_name: dict[str, list[str]] = {}
    iso_by_name: dict[str, list[str]] = {}
    for r in range(4, app.max_row + 1):
        name = app.cell(row=r, column=3).value
        nist = app.cell(row=r, column=6).value
        iso = app.cell(row=r, column=7).value
        if not name:
            continue
        name = str(name).strip()
        if nist:
            nist_by_name[name] = NIST_MAP.get(str(nist).strip(), [])
        if iso:
            iso_by_name[name] = [
                s.strip() for s in str(iso).replace(" / ", ",").split(",") if s.strip()
            ]

    kpis: list[dict[str, Any]] = []
    for r in range(3, ref.max_row + 1):
        num = ref.cell(row=r, column=1).value
        if num is None:
            continue
        category = ref.cell(row=r, column=2).value
        level_raw = ref.cell(row=r, column=3).value
        name = ref.cell(row=r, column=4).value
        max_score = ref.cell(row=r, column=5).value
        weightage = ref.cell(row=r, column=6).value
        tier_cols = [ref.cell(row=r, column=c).value for c in range(7, 12)]

        if not name:
            continue

        name = str(name).strip()
        tiers: list[TierOut] = []
        for i, t in enumerate(tier_cols, start=1):
            parsed = parse_tier(t)
            if parsed is None:
                continue
            tiers.append(
                TierOut(
                    tierLabel=parsed["tierLabel"],
                    scoreValue=parsed["scoreValue"],
                    condition=parsed["condition"],
                    tierOrder=i,
                )
            )

        level = LEVEL_MAP.get(
            str(level_raw).strip() if level_raw else "",
            "PROCESS",
        )
        # Category override: spreadsheet groups company-level KPIs
        # under TECHNOLOGY / MGMT REPORTING, but their app-level is COMPANY.
        if category and "MGMT" in str(category).upper() or (
            category and str(category).upper() == "TECHNOLOGY"
        ):
            level = "COMPANY"

        input_type = infer_input_type([t.tierLabel for t in tiers])

        kpis.append(
            asdict(
                KpiOut(
                    kpiName=name,
                    level=level,
                    category=str(category or "").strip(),
                    maxScore=int(max_score or 0),
                    weightage=float(weightage or 1.0),
                    inputType=input_type,
                    infoText=INFO_TEXT.get(
                        name, f"{name}. Provide the most recent measurement."
                    ),
                    nistControlIds=nist_by_name.get(name, []),
                    isoControlIds=iso_by_name.get(name, []),
                    tiers=tiers,
                )
            )
        )

    OUT.write_text(json.dumps(kpis, indent=2, ensure_ascii=False))
    print(f"Wrote {len(kpis)} KPIs to {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
